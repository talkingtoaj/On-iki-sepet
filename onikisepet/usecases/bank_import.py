import csv
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import NamedTuple

from django.core.exceptions import ValidationError
from django.db import transaction as db_transaction

from onikisepet import messages as msg
from onikisepet.money_input import parse_localized_decimal
from onikisepet.models import (
    Account,
    BankStatementImport,
    BankStatementRow,
    Category,
    Transaction,
)
from onikisepet.usecases import approval
from onikisepet.validators import validate_bank_import_file_extension

TURKISH_CHAR_MAP = str.maketrans("ığüşöç", "igusoc")

DATE_ALIASES = {"date", "tarih"}
DESCRIPTION_ALIASES = {"description", "aciklama", "desc"}
AMOUNT_ALIASES = {"amount", "tutar", "miktar", "islem_tutari"}
CURRENCY_ALIASES = {"currency", "para_birimi", "doviz"}
ACCOUNT_ALIASES = {"account", "hesap", "hesap_adi"}
TRANSACTION_TYPE_ALIASES = {"hareket_tipi"}

COLUMN_ALIASES = {
    "date": DATE_ALIASES,
    "description": DESCRIPTION_ALIASES,
    "amount": AMOUNT_ALIASES,
    "currency": CURRENCY_ALIASES,
    "account": ACCOUNT_ALIASES,
    "transaction_type": TRANSACTION_TYPE_ALIASES,
}

STANDARD_REQUIRED_COLUMNS = ("date", "description", "amount", "currency", "account")
TURKISH_BANK_REQUIRED_COLUMNS = ("date", "description", "amount")
SAMPLE_CSV_FILENAME = "ornek-ekstre.csv"
SAMPLE_CSV_ACCOUNT_FALLBACK = "Garanti - Ana Gider"

DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y")


class ConfirmImportResult(NamedTuple):
    bank_import: BankStatementImport
    imported_count: int
    pending_count: int


def build_sample_csv_content(*, account_name=None):
    resolved_account = account_name or SAMPLE_CSV_ACCOUNT_FALLBACK
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(STANDARD_REQUIRED_COLUMNS)
    writer.writerow(
        [
            "2026-06-09",
            "Ornek odeme",
            "125.50",
            "TRY",
            resolved_account,
        ]
    )
    return output.getvalue()


def normalize_header(value):
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value).strip().casefold())
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.replace(" ", "_").translate(TURKISH_CHAR_MAP)
    return text


def normalized_alias_set(aliases):
    return {normalize_header(alias) for alias in aliases}


def map_headers(raw_headers, required_columns=STANDARD_REQUIRED_COLUMNS):
    normalized_headers = [normalize_header(header) for header in raw_headers]
    mapped = {}

    for canonical_name, aliases in COLUMN_ALIASES.items():
        alias_set = normalized_alias_set(aliases)
        for index, header in enumerate(normalized_headers):
            if header in alias_set:
                mapped[canonical_name] = index
                break

    missing_columns = sorted(
        column for column in required_columns if column not in mapped
    )
    if missing_columns:
        raise ValidationError(
            msg.BANK_IMPORT_MISSING_COLUMNS.format(
                columns=", ".join(missing_columns),
            )
        )

    return mapped


def parse_date_value(value):
    if value is None:
        raise ValidationError("Tarih boş olamaz.")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        raise ValidationError("Tarih boş olamaz.")

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    raise ValidationError(f"Geçersiz tarih: {text}")


def parse_amount_value(value):
    amount = parse_localized_decimal(value)

    if amount == Decimal("0"):
        raise ValidationError("Tutar 0 olamaz.")

    return abs(amount)


def parse_currency_value(value):
    if value is None:
        raise ValidationError("Para birimi boş olamaz.")

    currency = str(value).strip().upper()
    valid_currencies = {choice for choice, _ in Account.Currency.choices}
    if currency not in valid_currencies:
        raise ValidationError(f"Geçersiz para birimi: {currency}")
    return currency


def resolve_account(account_name, accounts_by_name):
    if account_name is None:
        raise ValidationError("Hesap boş olamaz.")

    name = str(account_name).strip()
    if not name:
        raise ValidationError("Hesap boş olamaz.")

    account = accounts_by_name.get(name.lower())
    if account is None:
        raise ValidationError(f"Hesap bulunamadı: {name}")
    return account


def build_accounts_lookup():
    accounts = Account.objects.filter(is_active=True)
    return {account.name.lower(): account for account in accounts}


def row_values_from_mapping(values, header_map):
    return {
        key: values[index] if index < len(values) else None
        for key, index in header_map.items()
    }


def enrich_row_description(row_values):
    transaction_type = str(row_values.get("transaction_type") or "").strip()
    description = str(row_values.get("description") or "").strip()
    if transaction_type and description:
        return f"{transaction_type} — {description}"
    return transaction_type or description


def classification_from_hareket_tipi(raw_value):
    normalized = normalize_header(raw_value)
    if not normalized:
        return {}

    rules = (
        (("gider_transfer", "gidertransfer"), Transaction.TransactionType.TRANSFER, False),
        (("gelir_transfer", "gelirtransfer"), Transaction.TransactionType.TRANSFER, True),
        (("gider", "harcama"), Transaction.TransactionType.EXPENSE, False),
        (("gelir",), Transaction.TransactionType.INCOME, False),
    )
    for aliases, transaction_type, is_incoming in rules:
        if normalized in aliases:
            return {
                "transaction_type": transaction_type,
                "is_incoming_transfer": is_incoming,
            }
    return {}


def apply_import_defaults(row_values, *, default_account, default_currency):
    row_values = dict(row_values)
    row_values["hareket_tipi"] = str(row_values.get("transaction_type") or "").strip()
    if default_account is not None and not row_values.get("account"):
        row_values["account"] = default_account.name
    if default_currency and not row_values.get("currency"):
        row_values["currency"] = default_currency
    row_values["description"] = enrich_row_description(row_values)
    return row_values


def is_blank_row(values):
    return not any(
        value not in (None, "") and str(value).strip()
        for value in values
    )


def rows_from_table(table, *, default_account=None, default_currency=None):
    if not table:
        return []

    try:
        header_map = map_headers(table[0], required_columns=STANDARD_REQUIRED_COLUMNS)
    except ValidationError:
        if default_account is None:
            raise
        header_map = map_headers(table[0], required_columns=TURKISH_BANK_REQUIRED_COLUMNS)

    rows = []
    for values in table[1:]:
        values = list(values)
        if is_blank_row(values):
            continue
        row_values = row_values_from_mapping(values, header_map)
        row_values = apply_import_defaults(
            row_values,
            default_account=default_account,
            default_currency=default_currency or default_account.currency,
        )
        rows.append(row_values)
    return rows


def parse_row_values(row_values, *, row_number, accounts_by_name):
    errors = []

    try:
        parsed_date = parse_date_value(row_values.get("date"))
    except ValidationError as exc:
        errors.append(exc.messages[0])

    try:
        parsed_amount = parse_amount_value(row_values.get("amount"))
    except ValidationError as exc:
        errors.append(exc.messages[0])

    try:
        parsed_currency = parse_currency_value(row_values.get("currency"))
    except ValidationError as exc:
        errors.append(exc.messages[0])

    try:
        parsed_account = resolve_account(row_values.get("account"), accounts_by_name)
    except ValidationError as exc:
        errors.append(exc.messages[0])

    description = str(row_values.get("description") or "").strip()
    hareket_tipi = str(
        row_values.get("hareket_tipi") or row_values.get("transaction_type") or ""
    ).strip()

    if errors:
        return {
            "row_number": row_number,
            "description": description,
            "hareket_tipi": hareket_tipi,
            "parse_error": "; ".join(errors),
        }

    if parsed_account.currency != parsed_currency:
        return {
            "row_number": row_number,
            "description": description,
            "hareket_tipi": hareket_tipi,
            "parse_error": (
                f"Hesap para birimi ({parsed_account.currency}) "
                f"satırdaki para birimi ({parsed_currency}) ile uyuşmuyor."
            ),
        }

    return {
        "row_number": row_number,
        "date": parsed_date,
        "description": description,
        "amount": parsed_amount,
        "currency": parsed_currency,
        "account": parsed_account,
        "hareket_tipi": hareket_tipi,
        "parse_error": "",
    }


def read_csv_rows(uploaded_file):
    content = uploaded_file.read()
    if isinstance(content, bytes):
        text = content.decode("utf-8-sig")
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    try:
        raw_headers = next(reader)
    except StopIteration as exc:
        raise ValidationError(msg.BANK_IMPORT_EMPTY_FILE) from exc

    header_map = map_headers(raw_headers)
    rows = []
    for row_number, values in enumerate(reader, start=2):
        if is_blank_row(values):
            continue
        rows.append(row_values_from_mapping(values, header_map))
    return rows


def read_xlsx_rows(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError(
            "Excel dosyaları için openpyxl paketi gereklidir."
        ) from exc

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration as exc:
        raise ValidationError(msg.BANK_IMPORT_EMPTY_FILE) from exc

    header_map = map_headers(raw_headers)
    rows = []
    row_number = 2
    for values in row_iter:
        values = list(values)
        if is_blank_row(values):
            continue
        rows.append(row_values_from_mapping(values, header_map))
        row_number += 1
    return rows


def read_pdf_rows(uploaded_file, *, default_account=None):
    try:
        import pdfplumber
    except ImportError as exc:
        raise ValidationError(
            "PDF dosyaları için pdfplumber paketi gereklidir."
        ) from exc

    if default_account is None:
        raise ValidationError(msg.BANK_IMPORT_PDF_REQUIRES_ACCOUNT)

    rows = []
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                try:
                    rows.extend(
                        rows_from_table(
                            table,
                            default_account=default_account,
                            default_currency=default_account.currency,
                        )
                    )
                except ValidationError:
                    continue

    if not rows:
        raise ValidationError(msg.BANK_IMPORT_PDF_NO_PARSEABLE_ROWS)

    return rows


def read_uploaded_rows(uploaded_file, *, default_account=None):
    extension = Path(uploaded_file.name).suffix.lower()
    if extension == ".csv":
        return read_csv_rows(uploaded_file)
    if extension == ".xlsx":
        return read_xlsx_rows(uploaded_file)
    if extension == ".pdf":
        return read_pdf_rows(uploaded_file, default_account=default_account)
    raise ValidationError(msg.UNSUPPORTED_BANK_IMPORT_FILE)


def default_classification_for_account(account):
    if account is None:
        return {}
    if account.account_purpose != Account.AccountPurpose.ONLINE_DONATION:
        return {}

    defaults = {"transaction_type": Transaction.TransactionType.INCOME}
    category = Category.objects.filter(
        name=msg.ONLINE_DONATION_IMPORT_CATEGORY_NAME,
        category_type=Category.CategoryType.INCOME,
        is_active=True,
    ).first()
    if category is not None:
        defaults["category"] = category
    return defaults


def create_import_from_upload(uploaded_file, user, *, default_account=None):
    validate_bank_import_file_extension(uploaded_file)
    uploaded_file.seek(0)
    raw_rows = read_uploaded_rows(uploaded_file, default_account=default_account)
    if not raw_rows:
        raise ValidationError(msg.BANK_IMPORT_EMPTY_FILE)

    accounts_by_name = build_accounts_lookup()
    parsed_rows = [
        parse_row_values(row_values, row_number=index, accounts_by_name=accounts_by_name)
        for index, row_values in enumerate(raw_rows, start=1)
    ]

    with db_transaction.atomic():
        bank_import = BankStatementImport.objects.create(
            uploaded_by=user,
            original_filename=uploaded_file.name,
            status=BankStatementImport.Status.PREVIEW,
        )
        rows_to_create = []
        for row_data in parsed_rows:
            classification = {}
            if not row_data.get("parse_error"):
                classification = classification_from_hareket_tipi(
                    row_data.get("hareket_tipi"),
                )
                if not classification.get("transaction_type"):
                    classification = default_classification_for_account(
                        row_data.get("account"),
                    )
            rows_to_create.append(
                BankStatementRow(
                    bank_statement_import=bank_import,
                    row_number=row_data["row_number"],
                    date=row_data.get("date"),
                    description=row_data.get("description", ""),
                    amount=row_data.get("amount"),
                    currency=row_data.get("currency", ""),
                    account=row_data.get("account"),
                    parse_error=row_data.get("parse_error", ""),
                    transaction_type=classification.get("transaction_type", ""),
                    category=classification.get("category"),
                    is_incoming_transfer=classification.get(
                        "is_incoming_transfer",
                        False,
                    ),
                )
            )
        BankStatementRow.objects.bulk_create(rows_to_create)

    return bank_import


def get_importable_rows(bank_import):
    return bank_import.rows.filter(
        parse_error="",
        is_skipped=False,
        transaction__isnull=True,
    )


def is_row_ready_to_import(row):
    if row.transaction_id or row.is_skipped or row.parse_error:
        return False
    return validate_row_for_confirmation(row) is None


def get_row_workflow_status(row):
    if row.transaction_id:
        return "saved"
    if row.parse_error:
        return "error"
    if row.is_skipped:
        return "skipped"
    if is_row_ready_to_import(row):
        return "ready"
    return "pending"


def order_rows_for_preview(rows):
    status_rank = {
        "error": 0,
        "pending": 1,
        "ready": 2,
        "skipped": 3,
        "saved": 4,
    }
    return sorted(
        rows,
        key=lambda row: (
            status_rank.get(get_row_workflow_status(row), 9),
            row.row_number,
        ),
    )


def count_pending_rows(rows):
    return sum(
        1
        for row in rows
        if get_row_workflow_status(row) == "pending"
    )


def count_error_rows(rows):
    return sum(1 for row in rows if get_row_workflow_status(row) == "error")


def validate_row_for_confirmation(row):
    if row.parse_error:
        return msg.BANK_IMPORT_ROW_INVALID.format(
            row_number=row.row_number,
            error=row.parse_error,
        )
    if row.is_skipped:
        return None
    if not row.transaction_type:
        return msg.BANK_IMPORT_ROW_REQUIRES_TYPE.format(row_number=row.row_number)
    if row.transaction_type in (
        Transaction.TransactionType.INCOME,
        Transaction.TransactionType.EXPENSE,
    ) and row.category is None:
        return msg.BANK_IMPORT_ROW_REQUIRES_CATEGORY.format(row_number=row.row_number)
    if (
        row.transaction_type == Transaction.TransactionType.TRANSFER
        and row.target_account is None
    ):
        return msg.BANK_IMPORT_ROW_REQUIRES_TARGET_ACCOUNT.format(
            row_number=row.row_number,
        )
    return None


def build_transaction_from_row(row, user):
    transaction_kwargs = {
        "date": row.date,
        "amount": row.amount,
        "currency": row.currency,
        "transaction_type": row.transaction_type,
        "payee": row.payee,
        "description": row.description,
        "created_by": user,
    }

    if row.transaction_type == Transaction.TransactionType.INCOME:
        transaction_kwargs["target_account"] = row.account
        transaction_kwargs["category"] = row.category
    elif row.transaction_type == Transaction.TransactionType.EXPENSE:
        transaction_kwargs["source_account"] = row.account
        transaction_kwargs["category"] = row.category
    elif row.is_incoming_transfer:
        transaction_kwargs["source_account"] = row.target_account
        transaction_kwargs["target_account"] = row.account
    else:
        transaction_kwargs["source_account"] = row.account
        transaction_kwargs["target_account"] = row.target_account

    transaction = Transaction(**transaction_kwargs)
    approval.apply_initial_approval(transaction, user)
    return transaction


def confirm_import(bank_import, user):
    if bank_import.status == BankStatementImport.Status.CONFIRMED:
        raise ValidationError(msg.BANK_IMPORT_ALREADY_CONFIRMED)

    rows = list(bank_import.rows.order_by("row_number"))
    ready_rows = [row for row in rows if is_row_ready_to_import(row)]
    if not ready_rows:
        raise ValidationError(msg.BANK_IMPORT_NOT_READY)

    pending_count = count_pending_rows(rows)

    with db_transaction.atomic():
        for row in ready_rows:
            transaction = build_transaction_from_row(row, user)
            transaction.save()
            row.transaction = transaction
            row.save(update_fields=["transaction"])

        if pending_count == 0:
            bank_import.status = BankStatementImport.Status.CONFIRMED
            bank_import.save(update_fields=["status"])

    return ConfirmImportResult(
        bank_import=bank_import,
        imported_count=len(ready_rows),
        pending_count=pending_count,
    )
